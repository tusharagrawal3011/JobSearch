"""Seed the companies table.

Priority order:
  1. If company_tracker.xlsx exists, load from it (authoritative — Tushar's own file).
  2. Otherwise fall back to the inline seed below (the 15 confirmed + 4 unverified
     companies enumerated in the build spec).

For every company whose ats_type is 'unverified' OR whose api_url is missing, the
ATS-detection utility runs to resolve/confirm the slug + api_url. Confirmed companies
carry a best-guess slug that detection verifies and corrects if wrong.
"""
from __future__ import annotations

from typing import Optional

from backend import config
from backend.ats import detector
from backend.db.database import get_conn, now_iso

# SAMPLE seed — a few well-known companies that publish public ATS boards, so a fresh
# clone runs out-of-the-box and you can see discovery working. These are EXAMPLES, not a
# curated target list — replace them with your own via the dashboard's "Add company" /
# "Scout an area" features or a company_tracker.xlsx. Slugs are best-guess hints that the
# ATS detector verifies/corrects at seed time.
# Format: (name, ats_type, slug_hint, stack_fit, location, priority, notes)
INLINE_SEED: list[tuple] = [
    ("DevRev", "greenhouse", "devrev", "backend", "Bengaluru/Remote", "medium", "Sample (Greenhouse)"),
    ("Ashby", "ashby", "ashby", "backend", "Remote", "medium", "Sample (Ashby)"),
    ("Vercel", "greenhouse", "vercel", "frontend/backend", "Remote", "medium", "Sample (Greenhouse)"),
    ("Sarvam", "ashby", "sarvam", "ai/backend", "Bengaluru", "medium", "Sample (Ashby)"),
    ("Postman", "greenhouse", "postman", "backend", "Bengaluru/Remote", "medium", "Sample (Greenhouse)"),
    ("Hasura", "greenhouse", "hasura", "backend", "Bengaluru/Remote", "medium", "Sample (Greenhouse)"),
]


def _load_from_xlsx(path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip().lower() if h else "" for h in rows[0]]

    def col(row, *names):
        for n in names:
            if n in header:
                return row[header.index(n)]
        return None

    out = []
    for row in rows[1:]:
        name = col(row, "name", "company", "company name")
        if not name:
            continue
        out.append({
            "name": str(name).strip(),
            "ats_type": (str(col(row, "ats_type", "ats") or "unverified").strip().lower()),
            "ats_slug": _s(col(row, "ats_slug", "slug")),
            "api_url": _s(col(row, "api_url", "url")),
            "stack_fit": _s(col(row, "stack_fit", "stack")),
            "location": _s(col(row, "location")),
            "priority": _s(col(row, "priority")),
            "notes": _s(col(row, "notes")),
        })
    return out


def _s(v) -> Optional[str]:
    return str(v).strip() if v is not None and str(v).strip() else None


def _rows_from_inline() -> list[dict]:
    return [
        {"name": n, "ats_type": t, "ats_slug": slug, "api_url": None,
         "stack_fit": fit, "location": loc, "priority": pri, "notes": notes}
        for (n, t, slug, fit, loc, pri, notes) in INLINE_SEED
    ]


def resolve_ats(row: dict, verify: bool) -> dict:
    """Fill/confirm ats_type, ats_slug, api_url via the detector when needed."""
    needs = row["ats_type"] in (None, "", "unverified", "workday_likely") or not row.get("api_url")
    if verify and needs:
        res = detector.detect_ats(row["name"], explicit_slug=row.get("ats_slug"),
                                  careers_url=row.get("careers_url"))
        row["ats_type"] = res.ats_type if res.ats_type != "unverified" else (
            "unverified" if row["ats_type"] in (None, "", "workday_likely") else row["ats_type"]
        )
        if res.ats_type != "unverified":
            row["ats_slug"] = res.ats_slug
            row["api_url"] = res.api_url
    # Build a canonical api_url for confirmed rows even without live verification.
    if not row.get("api_url") and row["ats_type"] in ("greenhouse", "lever", "ashby") and row.get("ats_slug"):
        templ = {"greenhouse": detector.GREENHOUSE, "lever": detector.LEVER, "ashby": detector.ASHBY}
        row["api_url"] = templ[row["ats_type"]].format(slug=row["ats_slug"])
    if row["ats_type"] in (None, "", "workday_likely"):
        row["ats_type"] = "unverified"
    return row


def seed_companies(verify: bool = True) -> int:
    """Insert seed companies. `verify=True` hits the network via the detector.
    Returns number of companies inserted (skips ones already present by name)."""
    rows = _load_from_xlsx(config.COMPANY_TRACKER_XLSX) if config.COMPANY_TRACKER_XLSX.exists() else _rows_from_inline()
    inserted = 0
    with get_conn() as conn:
        existing = {r["name"] for r in conn.execute("SELECT name FROM companies").fetchall()}
        for row in rows:
            if row["name"] in existing:
                continue
            row = resolve_ats(row, verify)
            conn.execute(
                """INSERT INTO companies
                   (name, ats_type, ats_slug, api_url, stack_fit, location, priority, notes, added_at)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (row["name"], row["ats_type"], row.get("ats_slug"), row.get("api_url"),
                 row.get("stack_fit"), row.get("location"), row.get("priority"),
                 row.get("notes"), now_iso()),
            )
            inserted += 1
    return inserted


def add_company(name: str, careers_url: str = "", verify: bool = True) -> dict:
    """Dashboard 'Add company' entrypoint: insert + auto-detect ATS immediately.
    Returns the resolved row so the UI can show the detected type back to Tushar."""
    note = f"Added via dashboard. Careers URL: {careers_url}" if careers_url else "Added via dashboard"
    row = {"name": name.strip(), "ats_type": "unverified", "ats_slug": None,
           "api_url": None, "stack_fit": None, "location": None, "priority": "medium",
           "careers_url": careers_url, "notes": note}
    row = resolve_ats(row, verify)
    with get_conn() as conn:
        cur = conn.execute(
            """INSERT INTO companies
               (name, ats_type, ats_slug, api_url, careers_url, stack_fit, location, priority, notes, added_at)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (row["name"], row["ats_type"], row.get("ats_slug"), row.get("api_url"),
             careers_url or None, None, None, row.get("priority"), row.get("notes"), now_iso()),
        )
        row["id"] = cur.lastrowid
    return row
